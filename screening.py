import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys

# ── CONFIG ──────────────────────────────────────────────
RT_TOL = 0.1   # допуск по RT (хвилини)
MZ_TOL = 0.3   # допуск по m/z
# ────────────────────────────────────────────────────────

def classify(filename):
    f = str(filename).lower()
    if "blank" in f:
        return "blank"
    for i in range(1, 10):
        if f.startswith(f"{i}_") or f.startswith(f"{i}_neg"):
            return f"sample_{i}"
    return "unknown"

def coarse_screen(group, rt_tol=RT_TOL, mz_tol=MZ_TOL):
    files = group["File"].unique()
    if len(files) < 2:
        return pd.DataFrame()
    rep1 = group[group["File"] == files[0]]
    rep2 = group[group["File"] == files[1]]
    confirmed = []
    for _, p1 in rep1.iterrows():
        for _, p2 in rep2.iterrows():
            if abs(p1["RT"] - p2["RT"]) <= rt_tol and \
               abs(p1["Base Peak"] - p2["Base Peak"]) <= mz_tol:
                confirmed.append({
                    "Group":      f"{p1['SampleType']}_{p1['Polarity']}",
                    "RT_mean":    round((p1["RT"] + p2["RT"]) / 2, 4),
                    "MZ_mean":    round((p1["Base Peak"] + p2["Base Peak"]) / 2, 2),
                    "Area_mean":  int((p1["Area"] + p2["Area"]) / 2),
                    "Polarity":   p1["Polarity"],
                    "SampleType": p1["SampleType"],
                    "Rep1_Label": p1["Label"],
                    "Rep2_Label": p2["Label"],
                    "Confirmed":  "Yes"
                })
    return pd.DataFrame(confirmed)

def run(input_file, output_file):
    df = pd.read_excel(input_file)
    df.columns = [c.strip() for c in df.columns]
    df = df.dropna(subset=["RT"])
    df["SampleType"] = df["File"].apply(classify)

    # Coarse screening
    coarse_rows = []
    for (stype, pol), grp in df.groupby(["SampleType", "Polarity"]):
        result = coarse_screen(grp)
        if not result.empty:
            coarse_rows.append(result)
    coarse_df = pd.concat(coarse_rows, ignore_index=True) if coarse_rows else pd.DataFrame()

    # Out-target screening
    blanks  = coarse_df[coarse_df["SampleType"] == "blank"]
    samples = coarse_df[coarse_df["SampleType"] != "blank"]

    out_rows = []
    for _, peak in samples.iterrows():
        match = blanks[
            (blanks["Polarity"] == peak["Polarity"]) &
            (abs(blanks["RT_mean"] - peak["RT_mean"]) <= RT_TOL) &
            (abs(blanks["MZ_mean"] - peak["MZ_mean"]) <= MZ_TOL)
        ]
        status = "Artifact" if not match.empty else "Real Compound"
        row = peak.to_dict()
        row["Status"] = status
        if not match.empty:
            row["Blank_RT"] = match.iloc[0]["RT_mean"]
            row["Blank_MZ"] = match.iloc[0]["MZ_mean"]
        out_rows.append(row)
    out_df = pd.DataFrame(out_rows)

    # Summary
    summary_rows = []
    for (stype, pol), grp in df.groupby(["SampleType", "Polarity"]):
        confirmed = len(coarse_df[
            (coarse_df["SampleType"] == stype) &
            (coarse_df["Polarity"] == pol)
        ])
        if stype != "blank":
            sub = out_df[(out_df["SampleType"] == stype) & (out_df["Polarity"] == pol)]
            artifacts = len(sub[sub["Status"] == "Artifact"])
            real      = len(sub[sub["Status"] == "Real Compound"])
        else:
            artifacts = real = "-"
        summary_rows.append({
            "Sample": stype, "Polarity": pol,
            "Total peaks": len(grp),
            "Confirmed": confirmed,
            "Artifacts": artifacts,
            "Real Compounds": real
        })
    summary_df = pd.DataFrame(summary_rows)

    # Write Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer,       sheet_name="Raw Data",           index=False)
        coarse_df.to_excel(writer, sheet_name="Coarse Screening",  index=False)
        out_df.to_excel(writer,   sheet_name="Out-Target",         index=False)
        summary_df.to_excel(writer, sheet_name="Summary",          index=False)

    print(f"✅ Збережено: {output_file}")
    print(summary_df.to_string(index=False))

if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "input.xlsx"
    out = inp.replace(".xlsx", "_screened.xlsx")
    run(inp, out)
