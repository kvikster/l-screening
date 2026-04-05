from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
import pandas as pd
import io
import traceback
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from .logic import process_peaks
from .colors import read_with_operator_colors

app = FastAPI(title="LC-MS Screening API")

# Enable CORS for SvelteKit
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.post("/api/screen")
async def screen_file(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")
    
    try:
        content = await file.read()

        # Discover best sheet via pandas (fast, no colour)
        sheets_dict = pd.read_excel(io.BytesIO(content), sheet_name=None)
        required = ["RT", "Base Peak", "Polarity", "File", "Area"]
        best_sheet_name = max(
            sheets_dict,
            key=lambda n: len([c for c in required if c in [str(x).strip() for x in sheets_dict[n].columns]])
        )
        best_matches = len([c for c in required if c in [str(x).strip() for x in sheets_dict[best_sheet_name].columns]])
        if best_matches < 3:
            raise ValueError(f"Could not find a valid data sheet. Required columns: {', '.join(required)}")

        # Re-read the best sheet with openpyxl to capture operator cell colours
        df = read_with_operator_colors(content, best_sheet_name)
        results_df, summary_df, results_list = process_peaks(df)
        
        spec = {
            "root": "dashboard-1",
            "elements": {
                "dashboard-1": {
                    "type": "Dashboard",
                    "props": {
                        "title": f"Results from {file.filename} (Sheet: {best_sheet_name})",
                        "summary": summary_df.to_dict(orient="records"),
                        "peaks": results_list[:100]
                    },
                    "children": []
                }
            }
        }
        
        return spec

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/export")
async def export_file(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")

    try:
        content = await file.read()

        sheets_dict = pd.read_excel(io.BytesIO(content), sheet_name=None)
        required = ["RT", "Base Peak", "Polarity", "File", "Area"]
        best_sheet_name = max(
            sheets_dict,
            key=lambda n: len([c for c in required if c in [str(x).strip() for x in sheets_dict[n].columns]])
        )

        df_raw = read_with_operator_colors(content, best_sheet_name)
        results_df, summary_df, _ = process_peaks(df_raw)

        # Drop internal columns not useful in export
        raw_export = df_raw.drop(columns=["operator_color", "operator_mark"], errors="ignore")

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            raw_export.to_excel(writer, sheet_name="Raw Data", index=False)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            if not results_df.empty:
                export_cols = [c for c in [
                    "Group", "RT_mean", "MZ_mean", "Area_mean",
                    "SampleType", "Polarity", "Status",
                    "Rep1_Label", "Rep2_Label", "Rep1_Mark", "Rep2_Mark",
                ] if c in results_df.columns]
                results_df[export_cols].to_excel(writer, sheet_name="Screened Peaks", index=False)

                # Style header rows
                wb = writer.book
                header_fill = PatternFill("solid", fgColor="1E40AF")
                header_font = Font(bold=True, color="FFFFFF")
                for sheet in wb.worksheets:
                    for cell in sheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    for col in sheet.columns:
                        max_len = max((len(str(c.value or "")) for c in col), default=8)
                        sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

                # Color-code Status column in Screened Peaks
                sp = wb["Screened Peaks"]
                status_col_idx = None
                for i, cell in enumerate(sp[1], start=1):
                    if cell.value == "Status":
                        status_col_idx = i
                        break
                if status_col_idx:
                    green = PatternFill("solid", fgColor="DCFCE7")
                    red   = PatternFill("solid", fgColor="FEE2E2")
                    for row in sp.iter_rows(min_row=2):
                        cell = row[status_col_idx - 1]
                        if cell.value == "Real Compound":
                            cell.fill = green
                        elif cell.value == "Artifact":
                            cell.fill = red

        output.seek(0)
        stem = file.filename.rsplit(".", 1)[0]
        download_name = f"{stem}_screened.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
def health():
    return {"status": "ok"}
