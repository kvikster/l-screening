import io
import json
import math
import traceback
from collections import OrderedDict
from hashlib import sha256
from html import escape

import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .colors import read_with_operator_colors
from .logic import build_screening_config, process_peaks

app = FastAPI(title="LC-MS Screening API")
_SCREEN_CACHE_MAX_ITEMS = 8
_SCREEN_CACHE: "OrderedDict[str, dict]" = OrderedDict()

# Enable CORS for SvelteKit
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def _select_best_sheet(content: bytes) -> str:
    sheets_dict = pd.read_excel(io.BytesIO(content), sheet_name=None)
    required = ["RT", "Polarity", "File", "Area"]
    best_sheet_name = max(
        sheets_dict,
        key=lambda n: len(
            [
                c
                for c in required
                if c in [str(x).strip() for x in sheets_dict[n].columns]
            ]
        ),
    )
    best_matches = len(
        [
            c
            for c in required
            if c in [str(x).strip() for x in sheets_dict[best_sheet_name].columns]
        ]
    )
    if best_matches < 3:
        raise ValueError(
            f"Could not find a valid data sheet. Required columns: {', '.join(required)}"
        )
    return best_sheet_name


def _resolved_config_dict(config_overrides: dict) -> dict:
    config = build_screening_config(config_overrides)
    return {
        "replicate_rt_tol": config.replicate_rt_tol,
        "replicate_mz_tol": config.replicate_mz_tol,
        "replicate_mz_mode": config.replicate_mz_mode,
        "blank_rt_tol": config.blank_rt_tol,
        "blank_mz_tol": config.blank_mz_tol,
        "blank_mz_mode": config.blank_mz_mode,
        "signal_to_blank_min": config.signal_to_blank_min,
        "cv_high_max": config.cv_high_max,
        "cv_moderate_max": config.cv_moderate_max,
    }


def _screen_cache_key(content: bytes, resolved_config: dict) -> str:
    digest = sha256(content).hexdigest()
    config_digest = sha256(
        json.dumps(resolved_config, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return f"{digest}:{config_digest}"


def _screen_cache_get(cache_key: str) -> dict | None:
    cached = _SCREEN_CACHE.get(cache_key)
    if cached is None:
        return None
    _SCREEN_CACHE.move_to_end(cache_key)
    return {
        "results_df": cached["results_df"].copy(deep=True),
        "summary_df": cached["summary_df"].copy(deep=True),
        "results_list": list(cached["results_list"]),
        "best_sheet_name": cached["best_sheet_name"],
        "resolved_config": dict(cached["resolved_config"]),
        "df_raw": cached["df_raw"].copy(deep=True),
    }


def _screen_cache_put(cache_key: str, payload: dict) -> None:
    _SCREEN_CACHE[cache_key] = {
        "results_df": payload["results_df"].copy(deep=True),
        "summary_df": payload["summary_df"].copy(deep=True),
        "results_list": list(payload["results_list"]),
        "best_sheet_name": payload["best_sheet_name"],
        "resolved_config": dict(payload["resolved_config"]),
        "df_raw": payload["df_raw"].copy(deep=True),
    }
    _SCREEN_CACHE.move_to_end(cache_key)
    while len(_SCREEN_CACHE) > _SCREEN_CACHE_MAX_ITEMS:
        _SCREEN_CACHE.popitem(last=False)


def _screen_content(content: bytes, config_overrides: dict):
    resolved_config = _resolved_config_dict(config_overrides)
    cache_key = _screen_cache_key(content, resolved_config)
    cached = _screen_cache_get(cache_key)
    if cached is not None:
        return (
            cached["results_df"],
            cached["summary_df"],
            cached["results_list"],
            cached["best_sheet_name"],
            cached["resolved_config"],
            cached["df_raw"],
            True,
        )

    best_sheet_name = _select_best_sheet(content)
    df = read_with_operator_colors(content, best_sheet_name)
    results_df, summary_df, results_list = process_peaks(df, config=resolved_config)
    payload = {
        "results_df": results_df,
        "summary_df": summary_df,
        "results_list": results_list,
        "best_sheet_name": best_sheet_name,
        "resolved_config": resolved_config,
        "df_raw": df,
    }
    _screen_cache_put(cache_key, payload)
    return (
        results_df,
        summary_df,
        results_list,
        best_sheet_name,
        resolved_config,
        df.copy(deep=True),
        False,
    )


def _json_safe(value):
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    if hasattr(value, "item") and callable(value.item):
        try:
            return _json_safe(value.item())
        except (ValueError, TypeError):
            return value
    return value


def _build_export_payload(
    *,
    results_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    resolved_config: dict,
    df_raw: pd.DataFrame,
    best_sheet_name: str,
    file_name: str,
) -> dict:
    raw_export = df_raw.drop(
        columns=["operator_color", "operator_mark"], errors="ignore"
    )
    summary_export = summary_df.copy()
    params_export = pd.DataFrame(
        [{"Parameter": key, "Value": value} for key, value in resolved_config.items()]
    )

    results_export = pd.DataFrame()
    if not results_df.empty:
        results_export = results_df.copy()
        results_export["Why"] = results_export["Why"].apply(
            lambda value: json.dumps(value, ensure_ascii=False, sort_keys=True)
        )
        export_cols = [
            c
            for c in [
                "Group",
                "RT_mean",
                "MZ_mean",
                "Area_mean",
                "AreaCVPct",
                "ReplicateCount",
                "ReplicateQuality",
                "ReplicateConfidenceScore",
                "MatchingMode",
                "ParallelMatch",
                "ParallelSampleCount",
                "ParallelSourceSamples",
                "BlankAreaMean",
                "AreaDifference",
                "SignalToBlankRatio",
                "ConfidenceScore",
                "SampleType",
                "Polarity",
                "Status",
                "Rep1_Label",
                "Rep2_Label",
                "Rep1_Mark",
                "Rep2_Mark",
                "ReplicateFiles",
                "Why",
            ]
            if c in results_export.columns
        ]
        for list_col in (
            "ReplicateFiles",
            "ReplicateLabels",
            "ReplicateMarks",
            "ReplicateColors",
        ):
            if list_col in results_export.columns:
                results_export[list_col] = results_export[list_col].apply(
                    lambda value: json.dumps(value, ensure_ascii=False)
                )
        results_export = results_export[export_cols]

    return {
        "raw_export": raw_export,
        "summary_export": summary_export,
        "params_export": params_export,
        "results_export": results_export,
        "parameters": dict(resolved_config),
        "metadata": {
            "fileName": file_name,
            "sheetName": best_sheet_name,
            "rawRowCount": int(len(raw_export)),
            "summaryRowCount": int(len(summary_export)),
            "screenedRowCount": int(len(results_export)),
        },
    }


def _export_cell_to_text(value) -> str:
    normalized = _json_safe(value)
    if normalized is None:
        return ""
    if isinstance(normalized, (dict, list)):
        return json.dumps(normalized, ensure_ascii=False, sort_keys=True)
    return str(normalized)


def _df_to_records(df: pd.DataFrame) -> list[dict]:
    raw_records = _json_safe(df.to_dict(orient="records"))
    if isinstance(raw_records, list):
        return [record for record in raw_records if isinstance(record, dict)]
    return []


def _render_html_table(title: str, records: list[dict]) -> str:
    if not records:
        return f'<section class="card"><h2>{escape(title)}</h2><p class="empty">No data</p></section>'

    columns: list[str] = []
    for record in records:
        for key in record.keys():
            if key not in columns:
                columns.append(key)

    has_status = "Status" in columns
    table_slug = "".join(ch.lower() if ch.isalnum() else "-" for ch in title).strip("-")
    if not table_slug:
        table_slug = "table"

    header_html = "".join(
        f'<th class="sortable" data-col-index="{idx}" title="Click to sort">{escape(str(column))}<span class="sort-indicator">↕</span></th>'
        for idx, column in enumerate(columns)
    )

    row_html_parts: list[str] = []
    for record in records:
        cell_parts: list[str] = []
        for column in columns:
            value_text = _export_cell_to_text(record.get(column))
            if column == "Status":
                status_norm = value_text.strip().lower()
                badge_class = "badge-neutral"
                if status_norm == "real compound":
                    badge_class = "badge-real"
                elif status_norm == "artifact":
                    badge_class = "badge-artifact"
                rendered = f'<span class="status-badge {badge_class}">{escape(value_text)}</span>'
            else:
                rendered = escape(value_text)
            cell_parts.append(f"<td>{rendered}</td>")
        row_html_parts.append(f"<tr>{''.join(cell_parts)}</tr>")
    rows_html = "".join(row_html_parts)

    status_filter = ""
    if has_status:
        status_filter = f"""
      <label class="table-filter-label" for="{escape(table_slug)}-status-filter">Status:</label>
      <select id="{escape(table_slug)}-status-filter" class="table-status-filter" data-table-id="{escape(table_slug)}">
        <option value="">All</option>
        <option value="Real Compound">Real Compound</option>
        <option value="Artifact">Artifact</option>
      </select>
"""

    return (
        f'<section class="card">'
        f'<div class="section-head"><h2>{escape(title)}</h2></div>'
        f'<div class="table-controls">'
        f'<label class="table-filter-label" for="{escape(table_slug)}-search">Search:</label>'
        f'<input id="{escape(table_slug)}-search" class="table-search" type="search" '
        f'placeholder="Type to filter rows..." data-table-id="{escape(table_slug)}" />'
        f"{status_filter}"
        f'<button type="button" class="table-download" data-table-id="{escape(table_slug)}">Download filtered CSV</button>'
        f'<button type="button" class="table-reset" data-table-id="{escape(table_slug)}">Reset filters</button>'
        f"</div>"
        f'<div class="table-wrap"><table data-table-id="{escape(table_slug)}" data-sort-index="" data-sort-order="">'
        f"<thead><tr>{header_html}</tr></thead>"
        f"<tbody>{rows_html}</tbody></table></div></section>"
    )


def _status_metrics(records: list[dict]) -> dict:
    total = len(records)
    real_compounds = 0
    artifacts = 0

    for record in records:
        status = str(record.get("Status") or "").strip().lower()
        if status == "real compound":
            real_compounds += 1
        elif status == "artifact":
            artifacts += 1

    unknown = max(total - real_compounds - artifacts, 0)
    return {
        "total": total,
        "real_compounds": real_compounds,
        "artifacts": artifacts,
        "unknown": unknown,
    }


def _render_offline_html(payload: dict) -> str:
    metadata = payload["metadata"]
    parameters = payload["parameters"]

    summary_records = _df_to_records(payload["summary_export"])
    screened_records = _df_to_records(payload["results_export"])
    raw_records = _df_to_records(payload["raw_export"])
    params_records = _df_to_records(payload["params_export"])

    table_sections = [
        ("Summary", summary_records),
        ("Screened Peaks", screened_records),
        ("Raw Data", raw_records),
        ("Parameters", params_records),
    ]
    rendered_tables = "".join(
        _render_html_table(title, records) for title, records in table_sections
    )
    parameter_items = "".join(
        f"<li><strong>{escape(str(key))}</strong>: {escape(_export_cell_to_text(value))}</li>"
        for key, value in parameters.items()
    )

    metrics = _status_metrics(screened_records)

    return f"""<!doctype html>
<html lang="uk">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>LC-MS Screening Export</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f8fafc;
      --card: #ffffff;
      --text: #0f172a;
      --muted: #475569;
      --line: #e2e8f0;
      --accent: #1d4ed8;
      --accent-soft: #dbeafe;
      --good-bg: #dcfce7;
      --good-fg: #14532d;
      --bad-bg: #fee2e2;
      --bad-fg: #7f1d1d;
      --neutral-bg: #e2e8f0;
      --neutral-fg: #334155;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--text);
    }}
    .container {{
      max-width: 1400px;
      margin: 0 auto;
      padding: 24px;
    }}
    .hero, .card {{
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: 16px;
      box-shadow: 0 1px 2px rgba(15, 23, 42, 0.04);
    }}
    .hero {{
      padding: 20px;
      margin-bottom: 18px;
    }}
    .hero-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 8px;
    }}
    .print-btn {{
      border: 1px solid var(--line);
      background: white;
      color: var(--text);
      border-radius: 999px;
      padding: 8px 12px;
      font-size: 0.85rem;
      font-weight: 600;
      cursor: pointer;
    }}
    .print-btn:hover {{
      background: #f1f5f9;
    }}
    h1 {{
      margin: 0;
      font-size: 1.35rem;
    }}
    .meta {{
      display: grid;
      gap: 6px;
      color: var(--muted);
      font-size: 0.95rem;
    }}
    .metrics-grid {{
      margin-top: 12px;
      display: grid;
      gap: 8px;
      grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    }}
    .metric-card {{
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 10px 12px;
      background: #f8fafc;
    }}
    .metric-label {{
      font-size: 0.78rem;
      text-transform: uppercase;
      letter-spacing: 0.06em;
      color: var(--muted);
      font-weight: 700;
    }}
    .metric-value {{
      margin-top: 4px;
      font-size: 1.2rem;
      font-weight: 800;
      color: var(--text);
    }}
    .metric-card.real {{
      background: #f0fdf4;
      border-color: #bbf7d0;
    }}
    .metric-card.artifact {{
      background: #fef2f2;
      border-color: #fecaca;
    }}
    .params {{
      margin-top: 10px;
      padding-left: 18px;
      color: var(--muted);
    }}
    .card {{
      padding: 16px;
      margin-bottom: 14px;
      break-inside: avoid;
    }}
    .section-head {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 8px;
    }}
    h2 {{
      margin: 0;
      font-size: 1.05rem;
      color: var(--accent);
    }}
    .table-controls {{
      display: flex;
      align-items: center;
      flex-wrap: wrap;
      gap: 8px;
      margin-bottom: 10px;
    }}
    .table-filter-label {{
      color: var(--muted);
      font-size: 0.85rem;
      font-weight: 600;
    }}
    .table-search, .table-status-filter {{
      border: 1px solid var(--line);
      background: white;
      color: var(--text);
      border-radius: 10px;
      padding: 6px 10px;
      font-size: 0.85rem;
      min-width: 180px;
    }}
    .table-reset, .table-download {{
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      border-radius: 10px;
      padding: 6px 10px;
      font-size: 0.85rem;
      font-weight: 600;
      cursor: pointer;
    }}
    .table-reset:hover, .table-download:hover {{
      background: #f8fafc;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 12px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      background: white;
      font-size: 0.9rem;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      border-right: 1px solid var(--line);
      text-align: left;
      padding: 8px 10px;
      vertical-align: top;
      white-space: nowrap;
    }}
    th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: #eff6ff;
      color: #1e3a8a;
      font-weight: 700;
      user-select: none;
      cursor: pointer;
    }}
    th .sort-indicator {{
      margin-left: 6px;
      font-size: 0.75rem;
      color: #64748b;
    }}
    th.sorted-asc .sort-indicator {{
      color: #1d4ed8;
      content: "↑";
    }}
    th.sorted-desc .sort-indicator {{
      color: #1d4ed8;
      content: "↓";
    }}
    tr:nth-child(even) td {{
      background: #fcfdff;
    }}
    .status-badge {{
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 2px 8px;
      font-size: 0.8rem;
      font-weight: 700;
      border: 1px solid transparent;
    }}
    .badge-real {{
      background: var(--good-bg);
      color: var(--good-fg);
      border-color: #86efac;
    }}
    .badge-artifact {{
      background: var(--bad-bg);
      color: var(--bad-fg);
      border-color: #fca5a5;
    }}
    .badge-neutral {{
      background: var(--neutral-bg);
      color: var(--neutral-fg);
      border-color: #cbd5e1;
    }}
    .empty {{
      margin: 0;
      color: var(--muted);
    }}

    @media print {{
      :root {{
        --bg: #ffffff;
      }}
      body {{
        background: white;
      }}
      .container {{
        max-width: none;
        padding: 0;
      }}
      .hero, .card {{
        box-shadow: none;
        border-color: #d1d5db;
      }}
      .table-controls,
      .print-btn {{
        display: none !important;
      }}
      .table-wrap {{
        overflow: visible;
      }}
      table {{
        font-size: 10px;
      }}
      th {{
        position: static;
      }}
      tr, td, th {{
        break-inside: avoid;
      }}
      @page {{
        size: A4 landscape;
        margin: 12mm;
      }}
    }}
  </style>
</head>
<body>
  <main class="container">
    <section class="hero">
      <div class="hero-head">
        <h1>LC-MS Screening — Offline Export</h1>
        <button class="print-btn" onclick="window.print()">Print / Save PDF</button>
      </div>
      <div class="meta">
        <div><strong>File:</strong> {escape(_export_cell_to_text(metadata.get("fileName")))}</div>
        <div><strong>Sheet:</strong> {escape(_export_cell_to_text(metadata.get("sheetName")))}</div>
        <div><strong>Rows:</strong> raw={escape(_export_cell_to_text(metadata.get("rawRowCount")))}; summary={escape(_export_cell_to_text(metadata.get("summaryRowCount")))}; screened={escape(_export_cell_to_text(metadata.get("screenedRowCount")))}</div>
      </div>

      <div class="metrics-grid">
        <div class="metric-card">
          <div class="metric-label">Screened Total</div>
          <div class="metric-value">{metrics["total"]}</div>
        </div>
        <div class="metric-card real">
          <div class="metric-label">Real Compound</div>
          <div class="metric-value">{metrics["real_compounds"]}</div>
        </div>
        <div class="metric-card artifact">
          <div class="metric-label">Artifact</div>
          <div class="metric-value">{metrics["artifacts"]}</div>
        </div>
        <div class="metric-card">
          <div class="metric-label">Unknown</div>
          <div class="metric-value">{metrics["unknown"]}</div>
        </div>
      </div>

      <ul class="params">{parameter_items}</ul>
    </section>
    {rendered_tables}
  </main>

  <script>
    (function () {{
      function normalize(value) {{
        return String(value || "").toLowerCase().trim();
      }}

      function toNumberIfPossible(value) {{
        var cleaned = String(value || "").replace(/,/g, "").trim();
        if (cleaned === "") return null;
        var num = Number(cleaned);
        return Number.isFinite(num) ? num : null;
      }}

      function getCellText(row, index) {{
        var cells = row.querySelectorAll("td");
        if (!cells || !cells[index]) return "";
        return String(cells[index].textContent || "").trim();
      }}

      function applyTableFilter(tableId) {{
        var table = document.querySelector('table[data-table-id="' + tableId + '"]');
        if (!table) return;

        var searchInput = document.querySelector('.table-search[data-table-id="' + tableId + '"]');
        var statusSelect = document.querySelector('.table-status-filter[data-table-id="' + tableId + '"]');

        var q = normalize(searchInput ? searchInput.value : "");
        var status = normalize(statusSelect ? statusSelect.value : "");
        var rows = table.querySelectorAll("tbody tr");

        rows.forEach(function (row) {{
          var text = normalize(row.textContent);
          var statusCell = row.querySelector("td .status-badge");
          var rowStatus = normalize(statusCell ? statusCell.textContent : "");
          var matchText = !q || text.indexOf(q) !== -1;
          var matchStatus = !status || rowStatus === status;
          row.style.display = matchText && matchStatus ? "" : "none";
        }});
      }}

      function resetTable(tableId) {{
        var table = document.querySelector('table[data-table-id="' + tableId + '"]');
        if (!table) return;

        var searchInput = document.querySelector('.table-search[data-table-id="' + tableId + '"]');
        var statusSelect = document.querySelector('.table-status-filter[data-table-id="' + tableId + '"]');
        if (searchInput) searchInput.value = "";
        if (statusSelect) statusSelect.value = "";

        var tbody = table.querySelector("tbody");
        var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
        rows.sort(function (a, b) {{
          var ai = Number(a.getAttribute("data-original-index") || 0);
          var bi = Number(b.getAttribute("data-original-index") || 0);
          return ai - bi;
        }});
        rows.forEach(function (row) {{ tbody.appendChild(row); }});

        table.setAttribute("data-sort-index", "");
        table.setAttribute("data-sort-order", "");
        table.querySelectorAll("th.sortable").forEach(function (th) {{
          th.classList.remove("sorted-asc");
          th.classList.remove("sorted-desc");
          var indicator = th.querySelector(".sort-indicator");
          if (indicator) indicator.textContent = "↕";
        }});

        applyTableFilter(tableId);
      }}

      function csvEscape(value) {{
        var text = String(value == null ? "" : value);
        if (text.indexOf('"') !== -1) {{
          text = text.replace(/"/g, '""');
        }}
        if (text.indexOf(",") !== -1 || text.indexOf("\\n") !== -1 || text.indexOf("\\r") !== -1) {{
          text = '"' + text + '"';
        }}
        return text;
      }}

      function downloadFilteredCsv(tableId) {{
        var table = document.querySelector('table[data-table-id="' + tableId + '"]');
        if (!table) return;

        var headerCells = Array.prototype.slice.call(table.querySelectorAll("thead th"));
        var headers = headerCells.map(function (th) {{
          var labelNode = th.childNodes && th.childNodes.length ? th.childNodes[0] : null;
          var label = labelNode && labelNode.textContent ? labelNode.textContent : th.textContent;
          return String(label || "").trim();
        }});

        var lines = [];
        lines.push(headers.map(csvEscape).join(","));

        var rows = Array.prototype.slice.call(table.querySelectorAll("tbody tr"));
        rows.forEach(function (row) {{
          if (row.style.display === "none") return;
          var values = Array.prototype.slice.call(row.querySelectorAll("td")).map(function (td) {{
            return String(td.textContent || "").trim();
          }});
          lines.push(values.map(csvEscape).join(","));
        }});

        var csv = "\ufeff" + lines.join("\\n");
        var blob = new Blob([csv], {{ type: "text/csv;charset=utf-8;" }});
        var url = URL.createObjectURL(blob);
        var a = document.createElement("a");
        a.href = url;
        a.download = tableId + "_filtered.csv";
        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);
      }}

      function sortTableByColumn(tableId, colIndex) {{
        var table = document.querySelector('table[data-table-id="' + tableId + '"]');
        if (!table) return;

        var tbody = table.querySelector("tbody");
        var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
        var currentIndex = table.getAttribute("data-sort-index");
        var currentOrder = table.getAttribute("data-sort-order") || "asc";
        var nextOrder = (String(currentIndex) === String(colIndex) && currentOrder === "asc") ? "desc" : "asc";

        rows.sort(function (a, b) {{
          var aText = getCellText(a, colIndex);
          var bText = getCellText(b, colIndex);
          var aNum = toNumberIfPossible(aText);
          var bNum = toNumberIfPossible(bText);

          var cmp = 0;
          if (aNum !== null && bNum !== null) {{
            cmp = aNum - bNum;
          }} else {{
            cmp = aText.localeCompare(bText, undefined, {{ sensitivity: "base", numeric: true }});
          }}

          return nextOrder === "asc" ? cmp : -cmp;
        }});

        rows.forEach(function (row) {{ tbody.appendChild(row); }});
        table.setAttribute("data-sort-index", String(colIndex));
        table.setAttribute("data-sort-order", nextOrder);

        table.querySelectorAll("th.sortable").forEach(function (th) {{
          th.classList.remove("sorted-asc");
          th.classList.remove("sorted-desc");
          var indicator = th.querySelector(".sort-indicator");
          if (indicator) indicator.textContent = "↕";
        }});

        var active = table.querySelector('th.sortable[data-col-index="' + colIndex + '"]');
        if (active) {{
          active.classList.add(nextOrder === "asc" ? "sorted-asc" : "sorted-desc");
          var activeIndicator = active.querySelector(".sort-indicator");
          if (activeIndicator) activeIndicator.textContent = nextOrder === "asc" ? "↑" : "↓";
        }}
      }}

      document.querySelectorAll("table[data-table-id]").forEach(function (table) {{
        var tbody = table.querySelector("tbody");
        if (!tbody) return;
        Array.prototype.slice.call(tbody.querySelectorAll("tr")).forEach(function (row, idx) {{
          row.setAttribute("data-original-index", String(idx));
        }});

        table.querySelectorAll("th.sortable").forEach(function (th) {{
          th.addEventListener("click", function () {{
            var tableId = table.getAttribute("data-table-id");
            var colIndex = Number(th.getAttribute("data-col-index") || "0");
            sortTableByColumn(tableId, colIndex);
            applyTableFilter(tableId);
          }});
        }});
      }});

      document.querySelectorAll(".table-search").forEach(function (input) {{
        input.addEventListener("input", function (event) {{
          var target = event.target;
          applyTableFilter(target.getAttribute("data-table-id"));
        }});
      }});

      document.querySelectorAll(".table-status-filter").forEach(function (select) {{
        select.addEventListener("change", function (event) {{
          var target = event.target;
          applyTableFilter(target.getAttribute("data-table-id"));
        }});
      }});

      document.querySelectorAll(".table-download").forEach(function (button) {{
        button.addEventListener("click", function (event) {{
          var target = event.target;
          downloadFilteredCsv(target.getAttribute("data-table-id"));
        }});
      }});

      document.querySelectorAll(".table-reset").forEach(function (button) {{
        button.addEventListener("click", function (event) {{
          var target = event.target;
          resetTable(target.getAttribute("data-table-id"));
        }});
      }});
    }})();
  </script>
</body>
</html>"""


@app.post("/api/screen")
async def screen_file(
    file: UploadFile = File(...),
    replicate_rt_tol: float = Form(0.1),
    replicate_mz_tol: float = Form(0.3),
    replicate_mz_mode: str = Form("da"),
    blank_rt_tol: float = Form(0.1),
    blank_mz_tol: float = Form(0.3),
    blank_mz_mode: str = Form("da"),
    signal_to_blank_min: float = Form(3.0),
):
    file_name = file.filename or "uploaded.xlsx"
    if not file_name.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload an Excel file."
        )

    try:
        content = await file.read()
        config_overrides = {
            "replicate_rt_tol": replicate_rt_tol,
            "replicate_mz_tol": replicate_mz_tol,
            "replicate_mz_mode": replicate_mz_mode,
            "blank_rt_tol": blank_rt_tol,
            "blank_mz_tol": blank_mz_tol,
            "blank_mz_mode": blank_mz_mode,
            "signal_to_blank_min": signal_to_blank_min,
        }
        (
            results_df,
            summary_df,
            results_list,
            best_sheet_name,
            resolved_config,
            _,
            cache_hit,
        ) = _screen_content(content, config_overrides)
        shown_peaks = min(len(results_list), 1000)

        spec = {
            "root": "dashboard-1",
            "elements": {
                "dashboard-1": {
                    "type": "Dashboard",
                    "props": {
                        "title": f"Results from {file_name} (Sheet: {best_sheet_name})",
                        "summary": _json_safe(summary_df.to_dict(orient="records")),
                        "peaks": _json_safe(results_list[:1000]),
                        "parameters": _json_safe(resolved_config),
                        "metadata": _json_safe(
                            {
                                "sheetName": best_sheet_name,
                                "screenedPeakCount": len(results_df),
                                "totalPeaks": len(results_list),
                                "shownPeaks": shown_peaks,
                                "truncated": len(results_list) > shown_peaks,
                                "cacheHit": cache_hit,
                            }
                        ),
                    },
                    "children": [],
                }
            },
        }

        return spec

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/export")
async def export_file(
    file: UploadFile = File(...),
    replicate_rt_tol: float = Form(0.1),
    replicate_mz_tol: float = Form(0.3),
    replicate_mz_mode: str = Form("da"),
    blank_rt_tol: float = Form(0.1),
    blank_mz_tol: float = Form(0.3),
    blank_mz_mode: str = Form("da"),
    signal_to_blank_min: float = Form(3.0),
):
    file_name = file.filename or "uploaded.xlsx"
    if not file_name.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload an Excel file."
        )

    try:
        content = await file.read()
        config_overrides = {
            "replicate_rt_tol": replicate_rt_tol,
            "replicate_mz_tol": replicate_mz_tol,
            "replicate_mz_mode": replicate_mz_mode,
            "blank_rt_tol": blank_rt_tol,
            "blank_mz_tol": blank_mz_tol,
            "blank_mz_mode": blank_mz_mode,
            "signal_to_blank_min": signal_to_blank_min,
        }
        results_df, summary_df, _, best_sheet_name, resolved_config, df_raw, _ = (
            _screen_content(content, config_overrides)
        )

        payload = _build_export_payload(
            results_df=results_df,
            summary_df=summary_df,
            resolved_config=resolved_config,
            df_raw=df_raw,
            best_sheet_name=best_sheet_name,
            file_name=file_name,
        )
        raw_export = payload["raw_export"]
        summary_export = payload["summary_export"]
        params_export = payload["params_export"]
        results_export = payload["results_export"]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            raw_export.to_excel(writer, sheet_name="Raw Data", index=False)
            summary_export.to_excel(writer, sheet_name="Summary", index=False)
            params_export.to_excel(writer, sheet_name="Parameters", index=False)
            if not results_export.empty:
                results_export.to_excel(
                    writer, sheet_name="Screened Peaks", index=False
                )

            # Style header rows
            wb = writer.book
            header_fill = PatternFill("solid", fgColor="1E40AF")
            header_font = Font(bold=True, color="FFFFFF")
            for sheet in wb.worksheets:
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for col in sheet.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=8)
                    sheet.column_dimensions[
                        get_column_letter(col[0].column)
                    ].width = min(max_len + 4, 60)

            if "Screened Peaks" in wb.sheetnames:
                sp = wb["Screened Peaks"]
                status_col_idx = None
                for i, cell in enumerate(sp[1], start=1):
                    if cell.value == "Status":
                        status_col_idx = i
                        break
                if status_col_idx:
                    green = PatternFill("solid", fgColor="DCFCE7")
                    red = PatternFill("solid", fgColor="FEE2E2")
                    for row in sp.iter_rows(min_row=2):
                        cell = row[status_col_idx - 1]
                        if cell.value == "Real Compound":
                            cell.fill = green
                        elif cell.value == "Artifact":
                            cell.fill = red

        output.seek(0)
        stem = file_name.rsplit(".", 1)[0]
        download_name = f"{stem}_screened.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/export/html")
async def export_file_html(
    file: UploadFile = File(...),
    replicate_rt_tol: float = Form(0.1),
    replicate_mz_tol: float = Form(0.3),
    replicate_mz_mode: str = Form("da"),
    blank_rt_tol: float = Form(0.1),
    blank_mz_tol: float = Form(0.3),
    blank_mz_mode: str = Form("da"),
    signal_to_blank_min: float = Form(3.0),
):
    file_name = file.filename or "uploaded.xlsx"
    if not file_name.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=400, detail="Invalid file type. Please upload an Excel file."
        )

    try:
        content = await file.read()
        config_overrides = {
            "replicate_rt_tol": replicate_rt_tol,
            "replicate_mz_tol": replicate_mz_tol,
            "replicate_mz_mode": replicate_mz_mode,
            "blank_rt_tol": blank_rt_tol,
            "blank_mz_tol": blank_mz_tol,
            "blank_mz_mode": blank_mz_mode,
            "signal_to_blank_min": signal_to_blank_min,
        }
        results_df, summary_df, _, best_sheet_name, resolved_config, df_raw, _ = (
            _screen_content(content, config_overrides)
        )

        payload = _build_export_payload(
            results_df=results_df,
            summary_df=summary_df,
            resolved_config=resolved_config,
            df_raw=df_raw,
            best_sheet_name=best_sheet_name,
            file_name=file_name,
        )
        html_content = _render_offline_html(payload)

        output = io.BytesIO(html_content.encode("utf-8"))
        stem = file_name.rsplit(".", 1)[0]
        download_name = f"{stem}_screened_offline.html"
        return StreamingResponse(
            output,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
def health():
    return {"status": "ok"}
