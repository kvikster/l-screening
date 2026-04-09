import openpyxl
import io
import pandas as pd

# Standard Excel 56-color indexed palette (entries relevant to operator marks)
_INDEXED_TO_HEX = {
    11: "00FF00",  # Bright Green  → blank_negative
    13: "FFFF00",  # Yellow        → sample_rep2
    14: "FF00FF",  # Magenta       → sample_rep1
    15: "00FFFF",  # Cyan          → blank_positive
}

COLOR_SEMANTICS = {
    "00FF00": "blank_negative",
    "FFFF00": "sample_rep2",
    "FF00FF": "sample_rep1",
    "00FFFF": "blank_positive",
}

def _row_color_hex(row: tuple) -> str | None:
    """Return 6-char hex for the first cell's indexed fill, or None if uncolored."""
    first_cell = row[0]
    fill = first_cell.fill
    if not fill or fill.fill_type == "none":
        return None
    fc = fill.fgColor
    if fc.type == "indexed":
        return _INDEXED_TO_HEX.get(fc.indexed)
    if fc.type == "rgb":
        rgb = fc.rgb  # e.g. "FFRRGGBB"
        return rgb[-6:] if len(rgb) >= 6 else None
    return None


def read_with_operator_colors(content: bytes, sheet_name: str) -> pd.DataFrame:
    """
    Load *sheet_name* from *content* (raw xlsx bytes) and attach two extra columns:
      - operator_color : 6-char hex string or None
      - operator_mark  : semantic label (blank_positive, blank_negative,
                         sample_rep1, sample_rep2) or None
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2):
        color_hex = _row_color_hex(row)
        record = {h: row[i].value for i, h in enumerate(headers) if h is not None}
        record["operator_color"] = color_hex
        record["operator_mark"] = COLOR_SEMANTICS.get(color_hex) if color_hex else None
        records.append(record)

    return pd.DataFrame(records)
